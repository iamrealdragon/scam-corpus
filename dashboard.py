"""인도네시아 온라인스캠센터 보도 코퍼스 구축 현황 대시보드.

data/corpus.jsonl이 단일 소스(기존 시드 120건 + 발견 확장 398건 = 518건, 전량 병합됨)다.
seeds/existing_120.csv는 참고용으로만 남겨둔다(더 이상 별도 카운트 대상 아님 — 병합 완료).
LLM 프리코딩(coding_input/coding_sheet_v1_precoded.xlsx)과 α 신뢰도 표본
(coding_input/alpha_sample.xlsx) 진행 상황도 함께 보여준다.
"""
import json
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st

st.set_page_config(
    page_title="온라인스캠센터 보도 코퍼스 현황",
    page_icon="🗞️",
    layout="wide",
)

ROOT = Path(__file__).resolve().parent

# 8개 1차 매체, 고정 순서·고정 색상(Okabe-Ito 색맹 안전 팔레트)
OUTLET_ORDER = ["ANTARA", "Detik", "CNN Indonesia", "Liputan6", "Tirto", "Kumparan", "CNBC Indonesia", "Suara"]
OUTLET_COLOR = {
    "ANTARA": "#E69F00",
    "Detik": "#56B4E9",
    "CNN Indonesia": "#009E73",
    "Liputan6": "#D8B70A",
    "Tirto": "#0072B2",
    "Kumparan": "#D55E00",
    "CNBC Indonesia": "#CC79A7",
    "Suara": "#666666",
}

CODING_SHEET_PATH = ROOT / "coding_input" / "coding_sheet_v1_precoded.xlsx"
ALPHA_SAMPLE_PATH = ROOT / "coding_input" / "alpha_sample.xlsx"
CODEBOOK_FLAT_PATH = ROOT / "config" / "codebook_flat.json"


def normalize_outlet(name: str) -> str:
    for base in OUTLET_ORDER:
        if isinstance(name, str) and name.startswith(base):
            return base
    return name


@st.cache_data(ttl=60)
def variable_column_map() -> dict:
    """codebook_flat.json 변수 순서 = CodingSheet_KR 열 순서(1-based)."""
    cb = json.loads(CODEBOOK_FLAT_PATH.read_text(encoding="utf-8"))
    return {v["variable_id"]: i + 1 for i, v in enumerate(cb["variables"])}


@st.cache_data(ttl=60)
def load_seed() -> pd.DataFrame:
    path = ROOT / "seeds" / "existing_120.csv"
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_csv(path, encoding="utf-8-sig")
    df["outlet_norm"] = df["매체"].apply(normalize_outlet)
    return df


@st.cache_data(ttl=60)
def load_corpus() -> pd.DataFrame:
    path = ROOT / "data" / "corpus.jsonl"
    if not path.exists():
        return pd.DataFrame()
    records = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    df = pd.DataFrame(records)
    df["outlet_norm"] = df["outlet"].apply(normalize_outlet)
    # search_query 값으로 유래 구분: seed_merge(기존 120) vs tag_page/websearch_site_query(발견 확장)
    df["origin"] = df["search_query"].apply(lambda x: "기존 시드" if x == "seed_merge" else "발견 확장")
    return df


@st.cache_data(ttl=60)
def load_precoding_status() -> dict:
    import openpyxl

    if not CODING_SHEET_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(CODING_SHEET_PATH, data_only=True)
    ws = wb["CodingSheet_KR"]
    col_map = variable_column_map()
    h04_col, h05_col = col_map["H04"], col_map["H05"]

    total = 0
    succeeded = 0
    h04_dist = {"높음": 0, "중간": 0, "낮음": 0}
    h04_label = {"1": "높음", "2": "중간", "3": "낮음"}
    for r in range(4, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        total += 1
        h05 = ws.cell(r, h05_col).value
        if h05 == "LLM-1":
            succeeded += 1
            h04_raw = ws.cell(r, h04_col).value
            code = str(h04_raw).split()[0] if h04_raw else None
            if code in h04_label:
                h04_dist[h04_label[code]] += 1

    n_failed = 0
    if "프리코딩_실패" in wb.sheetnames:
        ws_fail = wb["프리코딩_실패"]
        n_failed = max(ws_fail.max_row - 1, 0)

    return {"총_대상": total, "성공": succeeded, "실패": n_failed, "H04_분포": h04_dist}


@st.cache_data(ttl=60)
def load_alpha_sample_status() -> pd.DataFrame:
    import openpyxl

    if not ALPHA_SAMPLE_PATH.exists():
        return pd.DataFrame()
    wb = openpyxl.load_workbook(ALPHA_SAMPLE_PATH, data_only=True)
    if "정답지_LLM사전코딩" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["정답지_LLM사전코딩"]
    col_map = variable_column_map()
    a02_col, h04_col = col_map["A02"], col_map["H04"]
    h04_label = {"1": "높음", "2": "중간", "3": "낮음"}

    rows = []
    for r in range(4, ws.max_row + 1):
        aid = ws.cell(r, 1).value
        if not aid:
            continue
        a02_raw = ws.cell(r, a02_col).value
        outlet = str(a02_raw).split(" ", 1)[1] if a02_raw and " " in str(a02_raw) else str(a02_raw)
        h04_raw = ws.cell(r, h04_col).value
        code = str(h04_raw).split()[0] if h04_raw else None
        rows.append({"article_id": aid, "매체": normalize_outlet(outlet), "확신도": h04_label.get(code, "?")})
    return pd.DataFrame(rows)


seed_df = load_seed()
corpus_df = load_corpus()

PAGES = ["연구 소개", "수집 현황", "매체별 현황", "연도별 추이", "코딩 진행", "검토 대기", "코퍼스 탐색"]
page = st.sidebar.radio("페이지", PAGES)

st.sidebar.markdown("---")
precoding = load_precoding_status()
st.sidebar.caption(
    f"코퍼스 총 {len(corpus_df)}건 (시드 병합 {int((corpus_df['origin']=='기존 시드').sum()) if not corpus_df.empty else 0}·"
    f"발견 확장 {int((corpus_df['origin']=='발견 확장').sum()) if not corpus_df.empty else 0}) · "
    f"프리코딩 {precoding.get('성공', 0)}/{precoding.get('총_대상', 0)}"
)


# ── 연구 소개 ────────────────────────────────────────────────────────────────
if page == "연구 소개":
    st.title("🗞️ 인도네시아 언론 온라인스캠센터 보도 코퍼스")
    st.markdown(
        """
### 연구 개요
인도네시아 언론(ANTARA·Detik·CNN Indonesia·Liputan6·Tirto·Kumparan·CNBC Indonesia·Suara)의
동남아 온라인 스캠센터(캄보디아·미얀마) 관련 WNI(인도네시아 국민) 피해 보도를 연구 목적으로
수집하는 프로젝트.

- 분석단위: 개별 기사
- 이론 골격: Entman 프레이밍 + Iyengar 에피소드/주제 프레임 + 플랫폼 책임 가시화 + 강제범죄화
- 방법론 템플릿: Meghan Sobel(2015 FGC, 2016 Thai sex trafficking) — 아카이브 기반 목적표집 + Krippendorff's α

### 현재 단계
**기존 120건 시드 병합 완료 + 발견 확장 3회차 완료 + LLM 1차 프리코딩 완료.**
사람 2인 α 신뢰도 표본(83건) 추출 대기 — 프리코딩은 사람 코딩을 대체하지 않는다.
        """
    )
    included_auto = 0
    if not corpus_df.empty:
        included_auto = int(((corpus_df["relevance_flag"] == "포함") & (corpus_df["access_status"] == "본문추출_자동")).sum())
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("코퍼스 총계", f"{len(corpus_df)}건")
    c2.metric("코딩 후보", f"{included_auto}건")
    c3.metric("프리코딩 완료", f"{precoding.get('성공', 0)}건")
    c4.metric("α 표본", f"{len(load_alpha_sample_status())}건")


# ── 수집 현황 ────────────────────────────────────────────────────────────────
elif page == "수집 현황":
    st.title("수집 현황")

    if corpus_df.empty:
        st.info("코퍼스 데이터 없음.")
    else:
        st.subheader("유래별 건수")
        origin_counts = corpus_df["origin"].value_counts()
        c1, c2 = st.columns(2)
        c1.metric("기존 시드(병합)", int(origin_counts.get("기존 시드", 0)))
        c2.metric("발견 확장", int(origin_counts.get("발견 확장", 0)))

        st.subheader("access_status 분포")
        status_counts = corpus_df["access_status"].value_counts()
        fig0 = px.bar(
            status_counts.rename_axis("access_status").reset_index(name="건수"),
            x="access_status", y="건수", text="건수",
            color_discrete_sequence=["#0072B2"],
        )
        fig0.update_traces(textposition="outside")
        fig0.update_layout(yaxis_title="건수", xaxis_title=None, showlegend=False)
        st.plotly_chart(fig0, use_container_width=True)

        st.subheader("relevance_flag 분포")
        rel_counts = corpus_df["relevance_flag"].value_counts()
        c1, c2 = st.columns(2)
        c1.metric("관련성 '포함'", int(rel_counts.get("포함", 0)))
        c2.metric("검토 대상('review')", int(rel_counts.get("review", 0)))

        method_counts = corpus_df["extraction_method"].value_counts()
        fig2 = px.bar(
            method_counts.rename_axis("추출 방식").reset_index(name="건수"),
            x="추출 방식", y="건수", text="건수",
            color_discrete_sequence=["#009E73"],
        )
        fig2.update_traces(textposition="outside")
        fig2.update_layout(yaxis_title="건수", xaxis_title=None, showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)

        n_clusters = corpus_df[corpus_df.get("cluster_size", pd.Series(dtype=int)).fillna(1) > 1]
        n_cluster_ids = n_clusters["event_cluster"].nunique() if not n_clusters.empty else 0
        st.caption(f"사건군 후보 클러스터: {n_cluster_ids}개 ({len(n_clusters)}건) — '검토 대기' 페이지에서 확인")


# ── 매체별 현황 ──────────────────────────────────────────────────────────────
elif page == "매체별 현황":
    st.title("매체별 현황")

    if corpus_df.empty:
        st.info("코퍼스 데이터 없음.")
    else:
        by_outlet_origin = (
            corpus_df.groupby(["outlet_norm", "origin"]).size().reset_index(name="건수")
        )
        fig = px.bar(
            by_outlet_origin, x="outlet_norm", y="건수", color="origin", barmode="group", text="건수",
            category_orders={"outlet_norm": OUTLET_ORDER},
            color_discrete_map={"기존 시드": "#666666", "발견 확장": "#0072B2"},
            labels={"outlet_norm": "매체", "origin": "구분"},
        )
        fig.update_traces(textposition="outside")
        fig.update_layout(yaxis_title="건수", xaxis_title=None, legend_title=None)
        st.plotly_chart(fig, use_container_width=True)

        st.subheader("매체별 상세")
        pivot = corpus_df.pivot_table(index="outlet_norm", columns="origin", aggfunc="size", fill_value=0)
        pivot = pivot.reindex(OUTLET_ORDER).fillna(0).astype(int)
        pivot.index.name = "매체"
        pivot["합계"] = pivot.sum(axis=1)
        st.dataframe(pivot, use_container_width=True)

        st.subheader("매체별 관련성 판정 비율 (전체 코퍼스)")
        rel_by_outlet = (
            corpus_df.groupby(["outlet_norm", "relevance_flag"]).size().reset_index(name="건수")
        )
        fig3 = px.bar(
            rel_by_outlet, x="outlet_norm", y="건수", color="relevance_flag", barmode="stack",
            category_orders={"outlet_norm": OUTLET_ORDER},
            color_discrete_map={"포함": "#009E73", "review": "#D55E00"},
            labels={"outlet_norm": "매체"},
        )
        fig3.update_layout(xaxis_title=None, legend_title=None)
        st.plotly_chart(fig3, use_container_width=True)


# ── 연도별 추이 ──────────────────────────────────────────────────────────────
elif page == "연도별 추이":
    st.title("연도별 추이")

    if corpus_df.empty:
        st.info("코퍼스 데이터 없음.")
    else:
        valid = corpus_df[corpus_df["year"].fillna(0) >= 2020]
        trend = valid.groupby(["year", "origin"]).size().reset_index(name="건수")

        fig = px.line(
            trend, x="year", y="건수", color="origin", markers=True,
            color_discrete_map={"기존 시드": "#666666", "발견 확장": "#0072B2"},
            labels={"year": "연도", "origin": "구분"},
        )
        fig.update_layout(xaxis_title=None, yaxis_title="건수", legend_title=None)
        fig.update_xaxes(type="category")
        st.plotly_chart(fig, use_container_width=True)
        st.caption("2026년은 아직 연중 — 하반기 반영 전 수치임에 유의. 1970 등 파싱 오류 연도는 제외(검토 대기에서 확인).")


# ── 코딩 진행 ────────────────────────────────────────────────────────────────
elif page == "코딩 진행":
    st.title("코딩 진행")
    st.caption("LLM 1차 프리코딩(claude-sonnet-4-6, Batch API) + 사람 2인 α 신뢰도 표본 상태.")

    if not precoding:
        st.info("coding_input/coding_sheet_v1_precoded.xlsx 없음 — 프리코딩 미착수.")
    else:
        st.subheader("LLM 프리코딩")
        c1, c2, c3 = st.columns(3)
        c1.metric("코딩 후보", precoding["총_대상"])
        c2.metric("성공", precoding["성공"])
        c3.metric("실패(크레딧 부족 등)", precoding["실패"])

        h04_df = pd.DataFrame(
            [{"확신도": k, "건수": v} for k, v in precoding["H04_분포"].items()]
        )
        fig = px.bar(
            h04_df, x="확신도", y="건수", text="건수",
            category_orders={"확신도": ["높음", "중간", "낮음"]},
            color_discrete_sequence=["#CC79A7"],
        )
        fig.update_traces(textposition="outside")
        fig.update_layout(yaxis_title="건수", xaxis_title=None, showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
        if precoding["H04_분포"].get("낮음", 0) == 0:
            st.warning(
                "H04(확신도) '낮음'이 0건이다 — LLM 과신 가능성. 사람 2인 α 결과와 대조해 확인 필요"
                "(STATUS.md 3번 항목 참고)."
            )

    st.subheader("α 신뢰도 표본")
    alpha_df = load_alpha_sample_status()
    if alpha_df.empty:
        st.info("coding_input/alpha_sample.xlsx 없음 — 표본 미추출.")
    else:
        st.metric("표본 크기", f"{len(alpha_df)}건")
        col1, col2 = st.columns(2)
        with col1:
            outlet_counts = alpha_df["매체"].value_counts().reindex(OUTLET_ORDER).fillna(0).astype(int)
            fig_o = px.bar(
                outlet_counts.rename_axis("매체").reset_index(name="건수"),
                x="매체", y="건수", text="건수",
                category_orders={"매체": OUTLET_ORDER},
                color_discrete_sequence=["#0072B2"],
            )
            fig_o.update_traces(textposition="outside")
            fig_o.update_layout(yaxis_title="건수", xaxis_title=None, showlegend=False)
            st.plotly_chart(fig_o, use_container_width=True)
        with col2:
            conf_counts = alpha_df["확신도"].value_counts()
            fig_c = px.bar(
                conf_counts.rename_axis("확신도").reset_index(name="건수"),
                x="확신도", y="건수", text="건수",
                category_orders={"확신도": ["높음", "중간", "낮음"]},
                color_discrete_sequence=["#CC79A7"],
            )
            fig_c.update_traces(textposition="outside")
            fig_c.update_layout(yaxis_title="건수", xaxis_title=None, showlegend=False)
            st.plotly_chart(fig_c, use_container_width=True)
        st.caption("다음 단계: coding_input/alpha_sample_coder1.xlsx / _coder2.xlsx로 사람 2인 독립 코딩 → Krippendorff's α 산출.")


# ── 검토 대기 ────────────────────────────────────────────────────────────────
elif page == "검토 대기":
    st.title("검토 대기")
    st.caption("자동 파이프라인이 확정하지 않고 사람 확인을 남겨둔 항목.")

    if corpus_df.empty:
        st.info("코퍼스 데이터 없음.")
    else:
        st.subheader("관련성 제외 후보 (relevance_flag = 'review')")
        review_df = corpus_df[corpus_df["relevance_flag"] == "review"][
            ["article_id", "outlet_norm", "title", "url"]
        ].rename(columns={"outlet_norm": "매체"})
        st.dataframe(review_df, use_container_width=True, hide_index=True)

        st.subheader("사건군 후보 클러스터 (AUTO- 상태, 사람 확정 전)")
        if "cluster_size" in corpus_df.columns:
            clustered = corpus_df[corpus_df["cluster_size"].fillna(1) > 1].sort_values(
                "event_cluster"
            )
            for cid, group in clustered.groupby("event_cluster"):
                with st.expander(f"{cid} — {len(group)}건"):
                    st.dataframe(
                        group[["article_id", "outlet_norm", "title", "pub_date", "url"]].rename(
                            columns={"outlet_norm": "매체"}
                        ),
                        use_container_width=True, hide_index=True,
                    )
        else:
            st.info("사건군 후보 없음.")

        st.subheader("데이터 이상 항목")
        issues = corpus_df[(corpus_df["title"].isna()) | (corpus_df["year"].fillna(0) < 2020)]
        if issues.empty:
            st.write("이상 없음.")
        else:
            st.dataframe(
                issues[["article_id", "outlet_norm", "title", "pub_date", "url"]].rename(
                    columns={"outlet_norm": "매체"}
                ),
                use_container_width=True, hide_index=True,
            )

        st.subheader("byline 오탐 (알려진 이슈, STATUS.md 3번 참고)")
        st.caption("A유형(날짜/시각/WIB): TIR 6건 · C유형(매체명이 사람이름 자리): ANT/DET/CNN 12건 · "
                    "B유형(Facebook URL): CNN 50건 — 별도 정정 스크립트 미작성 상태.")


# ── 코퍼스 탐색 ──────────────────────────────────────────────────────────────
elif page == "코퍼스 탐색":
    st.title("코퍼스 탐색")

    if corpus_df.empty:
        st.info("코퍼스 데이터 없음.")
    else:
        col1, col2, col3 = st.columns(3)
        outlet_filter = col1.multiselect("매체", OUTLET_ORDER, default=OUTLET_ORDER)
        relevance_filter = col2.multiselect(
            "관련성", ["포함", "review"], default=["포함", "review"]
        )
        origin_filter = col3.multiselect("구분", ["기존 시드", "발견 확장"], default=["기존 시드", "발견 확장"])
        keyword = st.text_input("제목 검색어")

        filtered = corpus_df[
            corpus_df["outlet_norm"].isin(outlet_filter)
            & corpus_df["relevance_flag"].isin(relevance_filter)
            & corpus_df["origin"].isin(origin_filter)
        ]
        if keyword:
            filtered = filtered[filtered["title"].fillna("").str.contains(keyword, case=False)]

        st.caption(f"{len(filtered)}건 표시 중")
        for _, row in filtered.iterrows():
            with st.expander(f"[{row['article_id']}] {row['outlet_norm']} · {row['origin']} · {row['title'] or '(제목 없음)'}"):
                st.write(f"발행일: {row['pub_date']} · 단어수: {row['word_count']} · "
                         f"관련성: {row['relevance_flag']} · access_status: {row['access_status']} · "
                         f"추출방식: {row['extraction_method']}")
                st.write(f"[원문 링크]({row['url']})")
                body = row["body_text"] or ""
                st.text(body[:600] + ("..." if len(body) > 600 else ""))
