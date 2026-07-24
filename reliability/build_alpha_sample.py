"""프리코딩 검증 + 사람 2인 신뢰도(α) 표본 추출.

[검증] coding_input/coding_sheet_v1_precoded.xlsx의 성공 378건 대상:
  - H04(코더 확신도) 분포
  - B01~H06 51개 변수별 결측(공란) 재확인
  - H04='낮음'(3) article_id 목록

[표본추출] 매체(A02) × 확신도(H04) 교차 층화, H04='낮음'은 전수 우선 포함,
나머지는 잔여 셀에 비례배분 후 각 셀 내 무작위 추출(고정 시드로 재현 가능).
목표 표본 비율은 TARGET_RATIO(기본 0.22 → 378건 기준 약 83건, 20~25% 범위 내).

출력(전부 신규 파일, 기존 파일은 읽기만 함):
  - coding_input/alpha_sample.xlsx        ("코딩시트" B~H 공란 + "정답지_LLM사전코딩" 시트)
  - coding_input/alpha_sample_coder1.xlsx (코더1용, "코딩시트" 단일 시트)
  - coding_input/alpha_sample_coder2.xlsx (코더2용, "코딩시트" 단일 시트, coder1과 동일 내용)
"""
import json
import random
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from crawler.precode_llm import load_codebook_variables, variable_column_index_map

ROOT = Path(__file__).resolve().parent.parent
PRECODED_PATH = ROOT / "coding_input" / "coding_sheet_v1_precoded.xlsx"
OUT_DIR = ROOT / "coding_input"

TARGET_RATIO = 0.22
RANDOM_SEED = 42

H04_LABELS = {1: "높음", 2: "중간", 3: "낮음"}


def load_success_rows() -> list[dict]:
    wb = openpyxl.load_workbook(PRECODED_PATH, data_only=True)
    ws = wb["CodingSheet_KR"]
    col_map = variable_column_index_map()
    h04_col = col_map["H04"]
    h05_col = col_map["H05"]
    a02_col = col_map["A02"]

    rows = []
    for r in range(4, ws.max_row + 1):
        aid = ws.cell(r, 1).value
        if not aid:
            continue
        h05_val = ws.cell(r, h05_col).value
        if h05_val != "LLM-1":  # 실패건은 H05가 비어있음
            continue
        h04_raw = ws.cell(r, h04_col).value  # "2 중간" 형식
        h04_code = int(str(h04_raw).split()[0]) if h04_raw else None
        a02_raw = ws.cell(r, a02_col).value
        outlet = str(a02_raw).split(" ", 1)[1] if a02_raw and " " in str(a02_raw) else str(a02_raw)
        rows.append({"row": r, "article_id": aid, "h04": h04_code, "outlet": outlet})
    return rows, ws, col_map


def verify(rows: list[dict], ws, col_map, variables: list[dict]) -> dict:
    h04_dist = {"높음": 0, "중간": 0, "낮음": 0}
    for r in rows:
        h04_dist[H04_LABELS.get(r["h04"], "?")] += 1

    missing_by_var = {}
    for v in variables:
        vid = v["variable_id"]
        col = col_map[vid]
        n_missing = sum(1 for r in rows if not ws.cell(r["row"], col).value)
        if n_missing:
            missing_by_var[vid] = n_missing

    low_ids = [r["article_id"] for r in rows if r["h04"] == 3]

    return {
        "총_성공_건수": len(rows),
        "H04_분포": h04_dist,
        "B01_H06_결측_변수": missing_by_var if missing_by_var else "전부 0건",
        "H04_낮음_article_id_목록": low_ids,
    }


def stratified_sample(rows: list[dict]) -> list[dict]:
    random.seed(RANDOM_SEED)
    target_n = round(len(rows) * TARGET_RATIO)

    low = [r for r in rows if r["h04"] == 3]
    non_low = [r for r in rows if r["h04"] != 3]

    selected = list(low)  # 낮음 전수 우선 포함
    remaining_quota = max(target_n - len(selected), 0)

    # 매체 x 확신도 교차 셀별로 비례배분 (잔여 풀 = non_low)
    cells = defaultdict(list)
    for r in non_low:
        cells[(r["outlet"], r["h04"])].append(r)

    total_non_low = len(non_low)
    if total_non_low and remaining_quota:
        # 비례배분 (최대잔여법으로 반올림 오차 보정)
        raw_alloc = {k: (len(v) / total_non_low) * remaining_quota for k, v in cells.items()}
        alloc = {k: int(raw_alloc[k]) for k in cells}
        allocated_sum = sum(alloc.values())
        remainder = remaining_quota - allocated_sum
        # 소수부 큰 순으로 1씩 배분
        for k in sorted(cells, key=lambda k: raw_alloc[k] - int(raw_alloc[k]), reverse=True)[:remainder]:
            alloc[k] += 1

        for k, cell_rows in cells.items():
            n = min(alloc.get(k, 0), len(cell_rows))
            selected.extend(random.sample(cell_rows, n))

    return selected


def build_header(ws_source):
    title = ws_source.cell(1, 1).value
    headers = [ws_source.cell(3, c).value for c in range(1, 64)]
    return title, headers


def write_coding_sheet_only(path: Path, title, headers, sample_rows, ws_source):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "코딩시트"
    ws.cell(1, 1, title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=63)
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h)

    row_idx = 4
    for sr in sample_rows:
        for c in range(1, 13):  # A01~A12만 복사
            ws.cell(row_idx, c, ws_source.cell(sr["row"], c).value)
        # B01~H06(col 13~63)은 공란으로 남김
        row_idx += 1
    wb.save(path)


def write_alpha_sample_with_answer_key(path: Path, title, headers, sample_rows, ws_source):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "코딩시트"
    ws1.cell(1, 1, title)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=63)
    for c, h in enumerate(headers, start=1):
        ws1.cell(3, c, h)
    row_idx = 4
    for sr in sample_rows:
        for c in range(1, 13):
            ws1.cell(row_idx, c, ws_source.cell(sr["row"], c).value)
        row_idx += 1

    ws2 = wb.create_sheet("정답지_LLM사전코딩")
    ws2.cell(1, 1, title + " (LLM 사전코딩값 — 코더에게 노출 금지, 비교용)")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=63)
    for c, h in enumerate(headers, start=1):
        ws2.cell(3, c, h)
    row_idx = 4
    for sr in sample_rows:
        for c in range(1, 64):
            ws2.cell(row_idx, c, ws_source.cell(sr["row"], c).value)
        row_idx += 1

    wb.save(path)


def main():
    variables = load_codebook_variables()
    # 검증에는 H05/H06 제외 목록에서 다시 B~H 전체(51개)로 확장
    from crawler.precode_llm import EXCLUDED_FROM_LLM
    import json as _json
    cb = _json.loads((ROOT / "config" / "codebook_flat.json").read_text(encoding="utf-8"))
    all_bh_vars = [v for v in cb["variables"] if not v["variable_id"].startswith("A")]

    rows, ws, col_map = load_success_rows()
    report = verify(rows, ws, col_map, all_bh_vars)
    print("=== 검증 결과 ===")
    print(json.dumps(report, ensure_ascii=False, indent=2))

    sample_rows = stratified_sample(rows)
    print(f"\n표본 크기: {len(sample_rows)} / {len(rows)} ({len(sample_rows)/len(rows)*100:.1f}%)")

    by_outlet = defaultdict(int)
    by_conf = defaultdict(int)
    cross = defaultdict(int)
    for r in sample_rows:
        by_outlet[r["outlet"]] += 1
        by_conf[H04_LABELS.get(r["h04"], "?")] += 1
        cross[(r["outlet"], H04_LABELS.get(r["h04"], "?"))] += 1

    print("\n매체별 표본 건수:", dict(by_outlet))
    print("확신도별 표본 건수:", dict(by_conf))
    print("매체x확신도 교차:", {f"{k[0]}/{k[1]}": v for k, v in cross.items()})

    title, headers = build_header(ws)
    write_alpha_sample_with_answer_key(OUT_DIR / "alpha_sample.xlsx", title, headers, sample_rows, ws)
    write_coding_sheet_only(OUT_DIR / "alpha_sample_coder1.xlsx", title, headers, sample_rows, ws)
    write_coding_sheet_only(OUT_DIR / "alpha_sample_coder2.xlsx", title, headers, sample_rows, ws)

    print("\n출력 파일:")
    print(" ", OUT_DIR / "alpha_sample.xlsx")
    print(" ", OUT_DIR / "alpha_sample_coder1.xlsx")
    print(" ", OUT_DIR / "alpha_sample_coder2.xlsx")

    return report, sample_rows, by_outlet, by_conf


if __name__ == "__main__":
    main()
