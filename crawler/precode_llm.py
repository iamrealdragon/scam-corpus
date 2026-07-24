"""LLM 1차 프리코딩 (CLAUDE.md 5번). B01~H04(H05/H06 제외) 49개 변수를 Claude로 채운다.

⚠️ 이 스크립트는 사람 2인 신뢰도 코딩(Krippendorff's α)을 대체하지 않는다.
   STATUS.md/CLAUDE.md 6번 방법론 메모대로, 신뢰도의 1차 앵커는 여전히 사람 코더 2인의
   20~25% 표본 독립 코딩이다. 여기서 나오는 값은 그 사람 코딩을 위한 "초안"일 뿐이며,
   LLM-사람 일치도는 보조 지표로만 별도 보고한다(라벨 절감 효과 근거, α 대체 아님).

입력: coding_input/coding_sheet_v1.xlsx (A01~A12 채워진 418행) + data/corpus.jsonl(body_text)
코드북: config/codebook_flat.json(B~H 49개 변수의 조작적정의·코딩규칙·예시)을 시스템 프롬프트에
        그대로 포함해 코드북 원문 기준으로만 판단하게 한다. 코드북에 없는 값은 strict tool
        schema(enum)로 원천 차단한다.
연구소개_온라인스캠센터.docx의 코딩 원칙(기사 내 표현만 근거, 배경지식으로 추정 금지,
        피해자성/가담자성 혼재 시 '혼합'/'경계적 주체' 코딩)을 시스템 프롬프트에 명시한다.

실행 방식: Anthropic Message Batches API(client.messages.batches) 사용, model=claude-sonnet-4-6
        (사용자 명시 지정). 강제 tool_choice + strict:true로 구조화 출력을 강제한다.
        원문 그대로 logs/precode_log.jsonl에 저장(재현성·감사용).

출력: coding_input/coding_sheet_v1_precoded.xlsx (A01~A12 유지, B01~H06 채움,
      H05='LLM-1' 고정, H06=실행일 자동), 실패건은 '프리코딩_실패' 시트에 별도 기록.
"""
import json
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import anthropic
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

ROOT = Path(__file__).resolve().parent.parent
CODEBOOK_FLAT_PATH = ROOT / "config" / "codebook_flat.json"
CODING_SHEET_PATH = ROOT / "coding_input" / "coding_sheet_v1.xlsx"
CORPUS_PATH = ROOT / "data" / "corpus.jsonl"
OUT_PATH = ROOT / "coding_input" / "coding_sheet_v1_precoded.xlsx"
LOG_PATH = ROOT / "logs" / "precode_log.jsonl"
FAIL_LOG_PATH = ROOT / "logs" / "precode_failures.jsonl"

MODEL = "claude-sonnet-4-6"  # 사용자 명시 지정
CODER_ID = "LLM-1"

EXCLUDED_FROM_LLM = {"H05", "H06"}  # H05=코더ID 고정, H06=코딩일 자동기입 — LLM이 채우지 않음

RESEARCH_INTRO = """\
## 연구 배경
동남아시아(캄보디아·미얀마 등)에서 온라인 스캠센터가 확산되며, 여러 국가 출신 이주자들이
허위 해외취업 제안·SNS 구인광고·비공식 모집망을 통해 이들 조직에 편입되는 사례가 보고된다.
피해자는 고객서비스·컴퓨터운영자·전자상거래운영자·식당근무 등 합법적 해외 일자리를 약속받지만
현지 도착 후 온라인 사기·온라인 도박·투자사기·로맨스 스캠 등 디지털 범죄 수행을 강요받는
경우가 있다. 이는 단순 해외취업 사기가 아니라 플랫폼화된 모집 + 강제 디지털 노동 + 피해자·가담자
경계가 겹치는 초국적 착취 양식이다.

## 연구 목적
인도네시아 언론이 인도네시아 국민(WNI)의 동남아 온라인 스캠센터 편입 문제를 어떻게
재현하는지 분석한다: (1) 기사가 이 문제를 취업사기/인신매매/온라인범죄/강제노동/외교적
보호/범죄가담/정책실패 중 무엇으로 정의하는지, (2) 플랫폼(Facebook/Telegram/WhatsApp/구인광고 등)이
단순 모집 경로로만 등장하는지 아니면 위험한 모집 인프라·책임 주체로 제시되는지, (3) 언론이
책임을 피해자 개인/모집책·브로커/스캠조직/플랫폼/인도네시아 정부/수용국 정부/국제범죄망 중
어디에 귀인하는지.
"""

CODING_PRINCIPLES = """\
## 코더(당신)의 역할과 원칙 — 반드시 지킬 것
1. 각 기사를 하나의 분석단위로 보고, **기사에 실제로 나타난 표현과 정보만**을 근거로 코딩한다.
2. **기사 밖의 배경지식으로 피해자성, 가담자성, 불법성, 책임 주체를 추정하지 않는다.**
   기사에서 WNI를 "피해자"로 명시하는 경우와 "스캐머"/"피의자"/"전직 스캐머"/"평가 대상"으로
   부르는 경우를 명확히 구분하라.
3. 같은 기사 안에서 피해자성과 가담자성이 동시에 나타나면 '혼합' 또는 '경계적 주체'로 코딩한다
   (배경지식으로 어느 한쪽으로 단정하지 말 것).
4. 모집 과정이 언급된 경우 모집 경로·플랫폼명·약속 직무·약속된 임금/혜택·실제 노동 전환 여부를
   기사에 명시된 만큼만 구체적으로 기록한다.
5. 강제 디지털 노동 관련 지표(여권 압수·감금·장시간노동·폭력/고문·임금체불/삭감·채무/몸값·
   실적목표·성폭력·디지털노동통제)는 기사에 **명시된 것만** 1로 코딩한다.
6. 책임 귀인은 원인 책임과 해결 책임을 나눠 판단한다. 플랫폼이 등장하더라도 단순 경로인지,
   문제적 인프라인지, 규제 대상인지, 명시적 책임 주체인지 구분한다.
7. **코드북에 없는 값을 만들지 않는다.** 아래 코드북에 제시된 코드/라벨만 사용하라(도구
   스키마가 이를 강제한다 — 스키마 밖 값은 애초에 낼 수 없다).
8. H01(원문 인용구)은 **15단어 이내**로 짧게, 저작권 고려해 긴 본문 복제는 금지한다.
9. H03(근거 메모)에는 왜 그 코드를 선택했는지 1~2문장으로 남긴다. 애매하거나 다중 프레임이
   가능한 경우 특히 근거를 남긴다.
10. H04(코더 확신도)는 스스로 판단한다: 기사가 모호하거나 정보가 부족하거나 스니펫에 의존한
    판단이면 3(낮음)으로, 명확하면 1(높음)로 표시한다.
"""


def load_codebook_variables() -> list[dict]:
    cb = json.loads(CODEBOOK_FLAT_PATH.read_text(encoding="utf-8"))
    return [v for v in cb["variables"] if v["variable_id"] not in EXCLUDED_FROM_LLM
            and not v["variable_id"].startswith("A")]


def variable_column_index_map() -> dict:
    """codebook_flat.json variables 순서 = CodingSheet_KR 열 순서(1-based)."""
    cb = json.loads(CODEBOOK_FLAT_PATH.read_text(encoding="utf-8"))
    return {v["variable_id"]: i + 1 for i, v in enumerate(cb["variables"])}


def build_codebook_reference_text(variables: list[dict]) -> str:
    lines = ["## 코드북 원문 (config/codebook_flat.json — 이 정의·규칙·허용값만 근거로 삼을 것)\n"]
    for v in variables:
        lines.append(f"### {v['variable_id']} {v['name']} ({v['dimension']}, 유형: {v['type']})")
        lines.append(f"- 조작적 정의: {v['definition']}")
        lines.append(f"- 코딩 규칙: {v['coding_rule']}")
        lines.append(f"- 예시: {v['example']}")
        if "allowed_codes" in v:
            opts = "; ".join(
                f"{o['code']} {o['label']}" if o["code"] is not None else o["label"]
                for o in v["allowed_codes"]
            )
            lines.append(f"- 허용값: {opts}")
        else:
            lines.append(f"- 입력형식: {v.get('allowed_format', '자유기입')}")
        lines.append("")
    return "\n".join(lines)


def build_tool_schema(variables: list[dict]) -> dict:
    properties = {}
    required = []
    for v in variables:
        vid = v["variable_id"]
        required.append(vid)
        desc = f"{v['name']}. {v['definition']} 규칙: {v['coding_rule']}"
        if v["type"] in ("Categorical", "Binary") and "allowed_codes" in v:
            codes = [o["code"] for o in v["allowed_codes"]]
            properties[vid] = {"type": "integer", "enum": codes, "description": desc}
        elif v["type"] == "Multiple / Text" and "allowed_codes" in v:
            labels = [o["label"] for o in v["allowed_codes"]]
            properties[vid] = {
                "type": "array",
                "items": {"type": "string", "enum": labels},
                "description": desc,
            }
        else:  # Text 자유기입 (H01/H02/H03 등)
            properties[vid] = {"type": "string", "description": desc}
    return {
        "type": "object",
        "properties": properties,
        "required": required,
        "additionalProperties": False,
    }


def build_system_prompt(codebook_ref: str) -> str:
    return RESEARCH_INTRO + "\n" + CODING_PRINCIPLES + "\n" + codebook_ref


def load_corpus_bodies() -> dict:
    bodies = {}
    for line in CORPUS_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        bodies[r["article_id"]] = {"title": r.get("title"), "body_text": r.get("body_text")}
    return bodies


def load_coding_candidates() -> list[dict]:
    wb = openpyxl.load_workbook(CODING_SHEET_PATH)
    ws = wb["CodingSheet_KR"]
    rows = []
    for r in range(4, ws.max_row + 1):
        article_id = ws.cell(r, 1).value
        if article_id:
            rows.append({"row": r, "article_id": article_id})
    return rows


def build_user_message(article_id: str, title: str, body_text: str) -> str:
    return (
        f"article_id: {article_id}\n"
        f"제목: {title}\n"
        f"본문:\n{body_text}\n\n"
        "위 기사를 코드북 기준으로 코딩해 submit_coding 도구를 호출하라. "
        "코드북에 없는 값은 만들지 말고, 기사에 실제로 나타난 표현만 근거로 삼아라."
    )


def main(limit: int | None = None):
    client = anthropic.Anthropic()

    variables = load_codebook_variables()
    print(f"LLM이 채울 변수 수(H05/H06 제외): {len(variables)}")
    codebook_ref = build_codebook_reference_text(variables)
    system_prompt = build_system_prompt(codebook_ref)
    tool_schema = build_tool_schema(variables)

    bodies = load_corpus_bodies()
    candidates = load_coding_candidates()
    if limit:
        candidates = candidates[:limit]
    print(f"코딩 대상 행 수: {len(candidates)}")

    tool = {
        "name": "submit_coding",
        "description": "기사 1건에 대한 B01~H04 코딩 결과를 제출한다.",
        "input_schema": tool_schema,
        "strict": True,
    }

    requests = []
    skipped_no_body = []
    for cand in candidates:
        aid = cand["article_id"]
        body_rec = bodies.get(aid)
        if not body_rec or not body_rec.get("body_text"):
            skipped_no_body.append(aid)
            continue
        requests.append(
            anthropic.types.messages.batch_create_params.Request(
                custom_id=aid,
                params=anthropic.types.message_create_params.MessageCreateParamsNonStreaming(
                    model=MODEL,
                    max_tokens=4096,
                    system=[{"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral"}}],
                    tools=[tool],
                    tool_choice={"type": "tool", "name": "submit_coding"},
                    messages=[{
                        "role": "user",
                        "content": build_user_message(aid, body_rec.get("title") or "", body_rec["body_text"]),
                    }],
                ),
            )
        )

    if skipped_no_body:
        print(f"본문 없어 스킵: {len(skipped_no_body)}건 -> {skipped_no_body[:10]}...")

    print(f"배치 요청 {len(requests)}건 제출 중...")
    batch = client.messages.batches.create(requests=requests)
    print(f"배치 ID: {batch.id}, 상태: {batch.processing_status}")

    while True:
        batch = client.messages.batches.retrieve(batch.id)
        counts = batch.request_counts
        print(f"  상태={batch.processing_status} "
              f"processing={counts.processing} succeeded={counts.succeeded} "
              f"errored={counts.errored} canceled={counts.canceled} expired={counts.expired}")
        if batch.processing_status == "ended":
            break
        time.sleep(30)

    col_map = variable_column_index_map()
    wb = openpyxl.load_workbook(CODING_SHEET_PATH)
    ws = wb["CodingSheet_KR"]
    row_by_id = {c["article_id"]: c["row"] for c in candidates}

    stats = {"succeeded": 0, "errored": 0, "canceled": 0, "expired": 0, "parse_failed": 0}
    failures = []
    schema_violations = []
    h04_dist = {1: 0, 2: 0, 3: 0}
    missing_by_var = {v["variable_id"]: 0 for v in variables}
    today = datetime.now().strftime("%Y-%m-%d")

    with open(LOG_PATH, "a", encoding="utf-8") as logf:
        for result in client.messages.batches.results(batch.id):
            aid = result.custom_id
            logf.write(json.dumps({
                "custom_id": aid, "result_type": result.result.type,
                "raw": result.model_dump() if hasattr(result, "model_dump") else str(result),
                "logged_at": datetime.now(timezone.utc).isoformat(),
            }, ensure_ascii=False, default=str) + "\n")

            if result.result.type != "succeeded":
                stats[result.result.type] = stats.get(result.result.type, 0) + 1
                failures.append({"article_id": aid, "reason": result.result.type})
                continue

            msg = result.result.message
            tool_use = next((b for b in msg.content if b.type == "tool_use"), None)
            if not tool_use:
                stats["parse_failed"] += 1
                failures.append({"article_id": aid, "reason": "no_tool_use_block"})
                continue

            coding = tool_use.input
            row = row_by_id.get(aid)
            if row is None:
                stats["parse_failed"] += 1
                failures.append({"article_id": aid, "reason": "row_not_found"})
                continue

            for v in variables:
                vid = v["variable_id"]
                val = coding.get(vid)
                col = col_map[vid]
                if val is None or val == "" or val == []:
                    missing_by_var[vid] += 1
                    ws.cell(row, col, "")
                    continue
                if v["type"] in ("Categorical", "Binary") and "allowed_codes" in v:
                    valid_codes = {o["code"] for o in v["allowed_codes"]}
                    if val not in valid_codes:
                        schema_violations.append({"article_id": aid, "var": vid, "value": val})
                    label = next((o["label"] for o in v["allowed_codes"] if o["code"] == val), "")
                    ws.cell(row, col, f"{val} {label}")
                    if vid == "H04":
                        h04_dist[val] = h04_dist.get(val, 0) + 1
                elif v["type"] == "Multiple / Text" and "allowed_codes" in v:
                    valid_labels = {o["label"] for o in v["allowed_codes"]}
                    invalid = [x for x in val if x not in valid_labels]
                    if invalid:
                        schema_violations.append({"article_id": aid, "var": vid, "value": invalid})
                    ws.cell(row, col, "; ".join(val))
                else:
                    ws.cell(row, col, val)

            ws.cell(row, col_map["H05"], CODER_ID)
            ws.cell(row, col_map["H06"], today)
            stats["succeeded"] += 1

    OUT_PATH.parent.mkdir(exist_ok=True)
    wb.save(OUT_PATH)

    if failures:
        with open(FAIL_LOG_PATH, "w", encoding="utf-8") as f:
            for fail in failures:
                f.write(json.dumps(fail, ensure_ascii=False) + "\n")
        ws_fail = wb.create_sheet("프리코딩_실패") if "프리코딩_실패" not in wb.sheetnames else wb["프리코딩_실패"]
        ws_fail.append(["article_id", "사유"])
        for fail in failures:
            ws_fail.append([fail["article_id"], fail["reason"]])
        wb.save(OUT_PATH)

    report = {
        "총_요청": len(requests),
        "본문없어_스킵": len(skipped_no_body),
        "성공": stats["succeeded"],
        "실패_상세": {k: v for k, v in stats.items() if k != "succeeded"},
        "H04_확신도_분포": h04_dist,
        "변수별_결측_건수(성공건_기준)": {k: v for k, v in missing_by_var.items() if v > 0},
        "스키마_위반_건수": len(schema_violations),
        "스키마_위반_상세": schema_violations[:20],
        "출력_경로": str(OUT_PATH),
        "로그_경로": str(LOG_PATH),
        "실패_로그_경로": str(FAIL_LOG_PATH) if failures else None,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return report


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()
    main(limit=args.limit)
