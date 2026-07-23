"""corpus.jsonl -> LLM 프리코딩 입력용 coding_sheet_v1.xlsx 생성.

CodingSheet_KR(config/codebook.xlsx)의 63열 헤더(A01~H06)를 그대로 복사해 시트 본체를 만들고,
A01~A12만 corpus.jsonl에서 자동 매핑해 채운다. B~H(51개 변수)는 LLM/사람 코딩 영역이라
전부 공란으로 남긴다. 코드북에 없는 값을 만들지 않는다(CLAUDE.md 10번).

대상: relevance_flag=='포함' & access_status=='본문추출_자동'인 레코드만.
나머지는 '제외_레코드' 시트에 사유와 함께 남긴다(삭제하지 않음).

QA 컬럼(64열째부터, 코딩시트 본체 63열 오른쪽)은 자동화 과정에서 발견된 이상치를
사람이 검토하도록 표시하는 용도다:
  - 발행일 확인 필요: A04 정규화 실패(1970-01-01 포함)/누락
  - 매체명 미매핑: outlet_key가 코드북 A02의 8개 매체에 없어 9(기타)로 처리된 경우
  - 사건군 미확정(AUTO): event_cluster가 아직 AUTO- 상태(사람이 KH-XXX-YYYY로 확정 전)
  - 1차코딩 사건군: seeds 원본(인도네시아_..._코딩북_1차.xlsx)의 사건군 값을 참고용으로만 병기
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

ROOT = Path(__file__).resolve().parent.parent
CORPUS_PATH = ROOT / "data" / "corpus.jsonl"
CODEBOOK_PATH = ROOT / "config" / "codebook.xlsx"
LEGACY_CODING_PATH = Path(
    "/Users/user/Downloads/인도네시아_온라인스캠센터_언론보도_코딩북_1차.xlsx"
)
OUT_DIR = ROOT / "coding_input"
OUT_PATH = OUT_DIR / "coding_sheet_v1.xlsx"

# A02 매체명 코드값 (config/codebook_flat.json A02와 동일 — 코드북에 없는 값 생성 금지)
OUTLET_KEY_TO_CODE_LABEL = {
    "antara": (1, "ANTARA"),
    "detik": (2, "Detik"),
    "cnn_indonesia": (3, "CNN Indonesia"),
    "liputan6": (4, "Liputan6"),
    "tirto": (5, "Tirto"),
    "kumparan": (6, "Kumparan"),
    "cnbc_indonesia": (7, "CNBC Indonesia"),
    "suara": (8, "Suara"),
}

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

QA_HEADERS = ["발행일 확인 필요", "매체명 미매핑", "사건군 미확정(AUTO)", "1차코딩 사건군"]


def load_corpus() -> list[dict]:
    return [json.loads(l) for l in CORPUS_PATH.read_text(encoding="utf-8").splitlines() if l.strip()]


def load_legacy_event_clusters() -> dict:
    """seeds 1차 코딩 파일의 legacy ID -> 사건군 매핑 (참고용, 값은 절대 덮어쓰지 않음)."""
    if not LEGACY_CODING_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(LEGACY_CODING_PATH, data_only=True)
    ws = wb["전체기사"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx_id = header.index("ID")
    idx_cluster = header.index("사건군")
    mapping = {}
    for row in rows[1:]:
        if row[idx_id] and row[idx_cluster]:
            mapping[row[idx_id]] = row[idx_cluster]
    return mapping


def pub_date_valid(d) -> bool:
    return bool(d) and bool(DATE_RE.match(str(d))) and str(d) != "1970-01-01"


def load_header_rows():
    wb = openpyxl.load_workbook(CODEBOOK_PATH, data_only=True)
    ws = wb["CodingSheet_KR"]
    title = ws.cell(1, 1).value
    headers = [ws.cell(3, c).value for c in range(1, 64)]
    return title, headers


def build_a_row(rec: dict, legacy_clusters: dict) -> tuple[list, dict]:
    """레코드 하나를 A01~A12(12칸) + QA 4칸으로 변환."""
    outlet_key = rec.get("outlet_key")
    code_label = OUTLET_KEY_TO_CODE_LABEL.get(outlet_key)
    outlet_unmapped = code_label is None
    if code_label:
        a02 = f"{code_label[0]} {code_label[1]}"
    else:
        a02 = f"9 기타({outlet_key})"

    a03 = "5 팩트체크" if rec.get("corpus_type") == "팩트체크" else ""

    pub_date = rec.get("pub_date")
    date_ok = pub_date_valid(pub_date)
    a04 = pub_date if date_ok else ""
    a05 = pub_date[:4] if date_ok else ""

    event_cluster = rec.get("event_cluster") or ""
    is_auto_cluster = event_cluster.startswith("AUTO-")

    search_query = rec.get("search_query") or rec.get("search_query_source") or ""

    a_cols = [
        rec.get("article_id"),           # A01
        a02,                              # A02
        a03,                              # A03
        a04,                              # A04
        a05,                              # A05
        "",                               # A06 (사람/LLM 영역)
        rec.get("url"),                   # A07
        event_cluster,                    # A08
        "",                               # A09 (사람/LLM 영역, 자동 추정 금지)
        "",                               # A10 (사람/LLM 영역)
        rec.get("word_count"),            # A11
        search_query,                     # A12
    ]

    legacy_id = rec.get("legacy_article_id")
    legacy_cluster = legacy_clusters.get(legacy_id) if legacy_id else None

    qa = {
        "발행일 확인 필요": "" if date_ok else "Y",
        "매체명 미매핑": outlet_key if outlet_unmapped else "",
        "사건군 미확정(AUTO)": "Y" if is_auto_cluster else "",
        "1차코딩 사건군": f"1차코딩 사건군: {legacy_cluster}" if legacy_cluster else "",
    }
    return a_cols, qa


def main():
    OUT_DIR.mkdir(exist_ok=True)
    records = load_corpus()
    legacy_clusters = load_legacy_event_clusters()
    title, headers = load_header_rows()

    included = [
        r for r in records
        if r.get("relevance_flag") == "포함" and r.get("access_status") == "본문추출_자동"
    ]
    included_ids = {r["article_id"] for r in included}
    excluded = [r for r in records if r["article_id"] not in included_ids]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CodingSheet_KR"

    # 헤더 복사 (본체 63열 그대로) + QA 4열 추가
    ws.cell(1, 1, title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=63)
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h)
    for i, qa_header in enumerate(QA_HEADERS):
        ws.cell(3, 64 + i, qa_header)

    stats_no_date = []
    stats_unmapped_outlet = []
    stats_auto_cluster = 0

    row_idx = 4
    for rec in included:
        a_cols, qa = build_a_row(rec, legacy_clusters)
        for c, val in enumerate(a_cols, start=1):
            ws.cell(row_idx, c, val)
        for i, qa_header in enumerate(QA_HEADERS):
            ws.cell(row_idx, 64 + i, qa[qa_header])

        if qa["발행일 확인 필요"] == "Y":
            stats_no_date.append(rec["article_id"])
        if qa["매체명 미매핑"]:
            stats_unmapped_outlet.append((rec["article_id"], rec.get("outlet_key")))
        if qa["사건군 미확정(AUTO)"] == "Y":
            stats_auto_cluster += 1

        row_idx += 1

    n_data_rows = row_idx - 4

    # 제외_레코드 시트
    ws2 = wb.create_sheet("제외_레코드")
    ws2.append(["article_id", "outlet", "relevance_flag", "access_status", "url", "사유"])
    for rec in excluded:
        reasons = []
        if rec.get("relevance_flag") != "포함":
            reasons.append(f"relevance_flag={rec.get('relevance_flag')}")
        if rec.get("access_status") != "본문추출_자동":
            reasons.append(f"access_status={rec.get('access_status')}")
        ws2.append([
            rec.get("article_id"), rec.get("outlet"), rec.get("relevance_flag"),
            rec.get("access_status"), rec.get("url"), "; ".join(reasons),
        ])

    wb.save(OUT_PATH)

    report = {
        "총_레코드": len(records),
        "코딩_후보(포함): 시트에 채운 행": n_data_rows,
        "제외_레코드": len(excluded),
        "A04_발행일_확인필요_건수": len(stats_no_date),
        "A04_발행일_확인필요_목록": stats_no_date,
        "A02_매체명_미매핑_건수": len(stats_unmapped_outlet),
        "A02_매체명_미매핑_목록": stats_unmapped_outlet,
        "A08_사건군_AUTO_건수": stats_auto_cluster,
        "A08_사건군_AUTO_비율": f"{stats_auto_cluster}/{n_data_rows}",
        "출력_경로": str(OUT_PATH),
        "시트_행수": row_idx - 1,  # 헤더 포함
        "시트_열수": 63 + len(QA_HEADERS),
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return report


if __name__ == "__main__":
    main()
